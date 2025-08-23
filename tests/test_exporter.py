from pathlib import Path

import openpyxl
import pandas as pd
from task_planner import exporter


def test_export_excel_creates_file(tmp_path: Path) -> None:
    schedule_df = pd.DataFrame(
        [
            {
                "Datum": "2025-09-07",
                "Tegenstander": "Test",
                "Uit/Thuis": "Uit",
                "Rijden 1": "Jan",
            }
        ]
    )
    stats_df = pd.DataFrame(
        [{"Speler": "Jan", "Totaal taken": 1, "Km": 20.0, "Rijden count": 1}]
    )
    out_file = tmp_path / "out.xlsx"

    exporter.export_excel(
        schedule_df, stats_df, players=["Jan"], out_path=str(out_file)
    )

    wb = openpyxl.load_workbook(out_file)
    assert "Schema" in wb.sheetnames
    assert "Statistiek" in wb.sheetnames
    ws = wb["Schema"]
    assert ws.cell(2, 4).value == "Jan"
