import sys
from argparse import Namespace
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import task_planner.__main__ as main
from pytest import MonkeyPatch


def test_full_flow_with_templates(tmp_path: Path) -> None:
    args = SimpleNamespace(
        taken="templates/taken.csv",
        wedstrijden="templates/wedstrijden.csv",
        spelers="templates/spelers.csv",
        afstanden="templates/afstanden.csv",
        uitvoer=tmp_path / "schema.xlsx",
        seed=42,
    )
    out_file = main.run_with_args(Namespace(**args.__dict__))
    assert Path(out_file).exists()

    wb = openpyxl.load_workbook(out_file)
    assert "Schema" in wb.sheetnames
    assert "Statistiek" in wb.sheetnames


def test_main_invokes_run_with_args(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    out_file = tmp_path / "schema.xlsx"
    argv = [
        "prog",
        "--taken",
        "templates/taken.csv",
        "--wedstrijden",
        "templates/wedstrijden.csv",
        "--spelers",
        "templates/spelers.csv",
        "--afstanden",
        "templates/afstanden.csv",
        "--uitvoer",
        str(out_file),
    ]
    monkeypatch.setattr(sys, "argv", argv)

    main.main()

    wb = openpyxl.load_workbook(out_file)
    assert "Schema" in wb.sheetnames
    assert "Statistiek" in wb.sheetnames
