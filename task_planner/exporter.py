try:
    import pandas as pd
except ImportError as err:
    raise SystemExit("Installatie nodig: pip install pandas openpyxl") from err

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as err:
    raise SystemExit("Installatie nodig: pip install pandas openpyxl") from err

from .utils import hashed_color_hex, to_int


def autosize_columns(ws: Worksheet) -> None:
    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) if cell.value else 0) for cell in col
        )
        col_idx = col[0].column
        if col_idx is not None:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 40
            )


def export_excel(
    schedule_df: pd.DataFrame,
    stats_df: pd.DataFrame,
    players: list[str],
    out_path: str,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert isinstance(ws, Worksheet)
    ws.title = "Schema"

    headers = list(schedule_df.columns)
    ws.append(headers)

    fills = {
        p: PatternFill(
            start_color=hashed_color_hex(p),
            end_color=hashed_color_hex(p),
            fill_type="solid",
        )
        for p in players
    }
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = bold
        cell.alignment = center

    name_cells_columns = [
        i
        for i, h in enumerate(headers, start=1)
        if any(
            h.startswith(prefix)
            for prefix in ["Materialen", "Hesjes", "Rijden", "Fluiten", "Bar"]
        )
    ]

    for i, row in schedule_df.iterrows():
        for j, h in enumerate(headers, start=1):
            val = row[h]
            cell = ws.cell(row=to_int(i) + 2, column=j, value=val)
            if (
                j in name_cells_columns
                and isinstance(val, str)
                and val in fills
            ):
                cell.fill = fills[val]
            if j <= 3:
                cell.alignment = center

    autosize_columns(ws)

    ws2 = wb.create_sheet("Statistiek")
    ws2.append(list(stats_df.columns))
    for j in range(1, stats_df.shape[1] + 1):
        ws2.cell(row=1, column=j).font = bold
        ws2.cell(row=1, column=j).alignment = center
    for _, r in stats_df.iterrows():
        ws2.append(list(r.values))  # TODO: Might want to sort this list somehow
    autosize_columns(ws2)

    # TODO: Catch PermissionError (most likely have excel open with same file)
    wb.save(out_path)  # TODO: Might want to use Path object
